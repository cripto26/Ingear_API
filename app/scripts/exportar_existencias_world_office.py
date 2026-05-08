from __future__ import annotations

import argparse
import os
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import TYPE_CHECKING, Any, Iterable, Sequence

if TYPE_CHECKING:
    import pyodbc


DEFAULT_VIEW = "Vista_ExistenciasPorBodegas"
DEFAULT_DATABASE = "WData"
DEFAULT_DRIVER = "SQL Server"

PRODUCT_CODE_COLUMNS = (
    "CodigoInventario",
    "Codigo_Inventario",
    "CodigoProducto",
    "Codigo_Producto",
    "Referencia",
    "Codigo",
)
PRODUCT_NAME_COLUMNS = (
    "DescripcionInventario",
    "NombreInventario",
    "DescripcionProducto",
    "NombreProducto",
    "Descripcion",
    "Nombre",
    "Producto",
)
STOCK_COLUMNS = (
    "Existencia",
    "Existencias",
    "Stock",
    "Saldo",
    "Cantidad",
)
BODEGA_CODE_COLUMNS = (
    "Codigo_Bodega",
    "CodigoBodega",
    "BodegaCodigo",
    "CodBodega",
)
BODEGA_NAME_COLUMNS = (
    "Nombre_Bodega",
    "NombreBodega",
    "Bodega",
    "DescripcionBodega",
)


class ReportError(RuntimeError):
    pass


def load_environment() -> None:
    script_path = Path(__file__).resolve()
    project_root = script_path.parents[2]
    try:
        from dotenv import load_dotenv
    except ImportError:
        load_env_file(project_root / ".env")
        load_env_file(Path.cwd() / ".env")
        return

    load_dotenv(project_root / ".env")
    load_dotenv()


def load_env_file(path: Path) -> None:
    if not path.exists():
        return

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def env(name: str, default: str | None = None) -> str | None:
    value = os.getenv(name)
    if value is None or value == "":
        return default
    return value


def quote_identifier(identifier: str) -> str:
    return "[" + identifier.replace("]", "]]") + "]"


def quote_qualified_name(name: str) -> str:
    parts = [part.strip() for part in name.split(".") if part.strip()]
    if not parts:
        raise ReportError("El nombre de la vista no puede estar vacio.")
    return ".".join(quote_identifier(part) for part in parts)


def normalize_column_lookup(columns: Iterable[str]) -> dict[str, str]:
    return {column.casefold(): column for column in columns}


def pick_column(
    columns: Sequence[str],
    candidates: Sequence[str],
    *,
    required: bool = False,
    label: str = "columna",
) -> str | None:
    lookup = normalize_column_lookup(columns)
    for candidate in candidates:
        match = lookup.get(candidate.casefold())
        if match:
            return match

    if required:
        available = ", ".join(columns)
        expected = ", ".join(candidates)
        raise ReportError(
            f"No encontre la {label}. Busque una de: {expected}. "
            f"Columnas disponibles: {available}"
        )
    return None


def build_connection_string(
    *,
    driver: str,
    server: str,
    database: str,
    username: str,
    password: str,
    timeout: int,
) -> str:
    return (
        f"DRIVER={{{driver}}};"
        f"SERVER={server};"
        f"DATABASE={database};"
        f"UID={username};"
        f"PWD={password};"
        "TrustServerCertificate=yes;"
        f"Connection Timeout={timeout};"
    )


def connect(args: argparse.Namespace) -> "pyodbc.Connection":
    try:
        import pyodbc
    except ImportError as exc:
        raise ReportError(
            "Falta pyodbc. Instalalo con: python -m pip install pyodbc"
        ) from exc

    server = args.server or env("WORLD_OFFICE_SERVER")
    username = args.username or env("WORLD_OFFICE_USERNAME")
    password = args.password or env("WORLD_OFFICE_PASSWORD")
    database = args.database or env("WORLD_OFFICE_DATABASE", DEFAULT_DATABASE)
    driver = args.driver or env("WORLD_OFFICE_ODBC_DRIVER", DEFAULT_DRIVER)
    timeout = int(args.connection_timeout)

    missing = [
        name
        for name, value in (
            ("servidor", server),
            ("usuario", username),
            ("clave", password),
            ("base de datos", database),
            ("driver ODBC", driver),
        )
        if not value
    ]
    if missing:
        raise ReportError("Faltan datos de conexion: " + ", ".join(missing))

    connection_string = build_connection_string(
        driver=str(driver),
        server=str(server),
        database=str(database),
        username=str(username),
        password=str(password),
        timeout=timeout,
    )

    try:
        connection = pyodbc.connect(connection_string)
    except pyodbc.Error as exc:
        raise ReportError(f"No se pudo conectar a World Office: {exc}") from exc

    connection.timeout = int(args.query_timeout)
    return connection


def get_view_columns(cursor: "pyodbc.Cursor", view_name: str) -> list[str]:
    quoted_view = quote_qualified_name(view_name)
    try:
        cursor.execute(f"SELECT TOP 0 * FROM {quoted_view}")
    except Exception as exc:
        raise ReportError(
            f"No se pudo leer la vista {view_name}. Revisa el nombre o permisos."
        ) from exc

    return [column[0] for column in cursor.description or []]


def clean_cell_value(value: object) -> object:
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return int(value)
        return float(value)
    return value


def fetch_total_inventory(
    cursor: "pyodbc.Cursor",
    *,
    view_name: str,
    bodega: str | None,
) -> tuple[list[dict[str, object]], dict[str, str | None]]:
    columns = get_view_columns(cursor, view_name)
    code_column = pick_column(
        columns,
        PRODUCT_CODE_COLUMNS,
        required=True,
        label="columna de codigo de producto",
    )
    name_column = pick_column(columns, PRODUCT_NAME_COLUMNS)
    stock_column = pick_column(
        columns,
        STOCK_COLUMNS,
        required=True,
        label="columna de existencias",
    )
    bodega_column = pick_column(columns, BODEGA_CODE_COLUMNS)

    quoted_view = quote_qualified_name(view_name)
    code_expr = (
        f"LTRIM(RTRIM(CAST({quote_identifier(str(code_column))} AS NVARCHAR(255))))"
    )
    params: list[object] = []
    where = [f"{quote_identifier(str(code_column))} IS NOT NULL", f"{code_expr} <> ''"]

    if bodega:
        if not bodega_column:
            raise ReportError(
                "Se envio filtro de bodega, pero la vista no tiene columna de bodega reconocida."
            )
        bodega_expr = (
            f"LTRIM(RTRIM(CAST({quote_identifier(str(bodega_column))} AS NVARCHAR(255))))"
        )
        where.append(f"{bodega_expr} = ?")
        params.append(bodega)

    select_parts = [f"{code_expr} AS codigo_producto"]
    if name_column:
        select_parts.append(
            f"MAX(CAST({quote_identifier(str(name_column))} AS NVARCHAR(500))) AS producto"
        )
    select_parts.append(
        f"SUM(COALESCE({quote_identifier(str(stock_column))}, 0)) AS existencia"
    )

    query = f"""
        SELECT
            {", ".join(select_parts)}
        FROM {quoted_view}
        WHERE {" AND ".join(where)}
        GROUP BY {code_expr}
        ORDER BY codigo_producto
    """
    cursor.execute(query, params)

    rows: list[dict[str, object]] = []
    for raw_row in cursor.fetchall():
        index = 0
        row: dict[str, object] = {"Codigo producto": raw_row[index]}
        index += 1
        if name_column:
            row["Producto"] = raw_row[index]
            index += 1
        row["Existencia"] = clean_cell_value(raw_row[index])
        rows.append(row)

    metadata = {
        "code_column": code_column,
        "name_column": name_column,
        "stock_column": stock_column,
        "bodega_column": bodega_column,
    }
    return rows, metadata


def fetch_inventory_by_bodega(
    cursor: "pyodbc.Cursor",
    *,
    view_name: str,
    bodega: str | None,
) -> list[dict[str, object]]:
    columns = get_view_columns(cursor, view_name)
    code_column = pick_column(columns, PRODUCT_CODE_COLUMNS, required=True)
    name_column = pick_column(columns, PRODUCT_NAME_COLUMNS)
    stock_column = pick_column(columns, STOCK_COLUMNS, required=True)
    bodega_column = pick_column(columns, BODEGA_CODE_COLUMNS)
    bodega_name_column = pick_column(columns, BODEGA_NAME_COLUMNS)
    if not bodega_column:
        return []

    quoted_view = quote_qualified_name(view_name)
    code_expr = (
        f"LTRIM(RTRIM(CAST({quote_identifier(str(code_column))} AS NVARCHAR(255))))"
    )
    bodega_expr = (
        f"LTRIM(RTRIM(CAST({quote_identifier(str(bodega_column))} AS NVARCHAR(255))))"
    )
    params: list[object] = []
    where = [f"{quote_identifier(str(code_column))} IS NOT NULL", f"{code_expr} <> ''"]

    if bodega:
        where.append(f"{bodega_expr} = ?")
        params.append(bodega)

    select_parts = [f"{code_expr} AS codigo_producto"]
    if name_column:
        select_parts.append(
            f"MAX(CAST({quote_identifier(str(name_column))} AS NVARCHAR(500))) AS producto"
        )
    select_parts.append(f"{bodega_expr} AS codigo_bodega")
    if bodega_name_column:
        select_parts.append(
            f"MAX(CAST({quote_identifier(str(bodega_name_column))} AS NVARCHAR(500))) AS bodega"
        )
    select_parts.append(
        f"SUM(COALESCE({quote_identifier(str(stock_column))}, 0)) AS existencia"
    )

    group_by = [code_expr, bodega_expr]
    query = f"""
        SELECT
            {", ".join(select_parts)}
        FROM {quoted_view}
        WHERE {" AND ".join(where)}
        GROUP BY {", ".join(group_by)}
        ORDER BY codigo_producto, codigo_bodega
    """
    cursor.execute(query, params)

    rows: list[dict[str, object]] = []
    for raw_row in cursor.fetchall():
        index = 0
        row: dict[str, object] = {"Codigo producto": raw_row[index]}
        index += 1
        if name_column:
            row["Producto"] = raw_row[index]
            index += 1
        row["Codigo bodega"] = raw_row[index]
        index += 1
        if bodega_name_column:
            row["Bodega"] = raw_row[index]
            index += 1
        row["Existencia"] = clean_cell_value(raw_row[index])
        rows.append(row)
    return rows


def write_rows_as_table(
    workbook: Any,
    *,
    sheet_name: str,
    title: str,
    rows: Sequence[dict[str, object]],
    table_name: str,
) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    sheet = workbook.create_sheet(sheet_name)
    sheet.sheet_view.showGridLines = False

    headers = list(rows[0].keys()) if rows else ["Codigo producto", "Existencia"]
    sheet.append([title])
    sheet.append(headers)
    for row in rows:
        sheet.append([clean_cell_value(row.get(header)) for header in headers])

    title_range = f"A1:{get_column_letter(len(headers))}1"
    sheet.merge_cells(title_range)
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet["A1"].alignment = Alignment(horizontal="center")

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for cell in sheet[2]:
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top")
        row[-1].number_format = "#,##0.####"

    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{sheet.max_row}"

    if rows:
        ref = f"A2:{get_column_letter(len(headers))}{sheet.max_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        values = [str(cell.value or "") for cell in column_cells]
        width = min(max(max((len(value) for value in values), default=10) + 2, 12), 60)
        sheet.column_dimensions[column_letter].width = width


def create_workbook(
    *,
    total_rows: Sequence[dict[str, object]],
    detail_rows: Sequence[dict[str, object]],
    metadata: dict[str, str | None],
    args: argparse.Namespace,
) -> Any:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise ReportError(
            "Falta openpyxl. Instalalo con: python -m pip install openpyxl"
        ) from exc

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumen"
    summary.sheet_view.showGridLines = False

    total_existencias = sum(
        Decimal(str(row.get("Existencia") or 0)) for row in total_rows
    )
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary_rows = [
        ("Informe", "Existencias de productos"),
        ("Generado", generated_at),
        ("Servidor", args.server or env("WORLD_OFFICE_SERVER")),
        ("Base de datos", args.database or env("WORLD_OFFICE_DATABASE", DEFAULT_DATABASE)),
        ("Vista", args.view),
        ("Filtro bodega", args.bodega or "Todas"),
        ("Productos", len(total_rows)),
        ("Existencia total", clean_cell_value(total_existencias)),
        ("Columna codigo", metadata.get("code_column")),
        ("Columna producto", metadata.get("name_column") or "No disponible"),
        ("Columna existencia", metadata.get("stock_column")),
        ("Columna bodega", metadata.get("bodega_column") or "No disponible"),
    ]
    summary.append(["Campo", "Valor"])
    for item in summary_rows:
        summary.append(list(item))

    summary["A1"].font = Font(bold=True, color="FFFFFF")
    summary["B1"].font = Font(bold=True, color="FFFFFF")
    for cell in summary[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")

    for row in summary.iter_rows(min_row=2, max_row=summary.max_row):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(horizontal="left")
    summary["B8"].number_format = "#,##0.####"
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 52
    summary.freeze_panes = "A2"

    write_rows_as_table(
        workbook,
        sheet_name="Existencias",
        title="Existencias por producto",
        rows=total_rows,
        table_name="TablaExistencias",
    )

    if detail_rows:
        write_rows_as_table(
            workbook,
            sheet_name="Por bodega",
            title="Existencias por producto y bodega",
            rows=detail_rows,
            table_name="TablaExistenciasBodega",
        )

    return workbook


def resolve_output_path(output: str | None) -> Path:
    if output:
        return Path(output).expanduser().resolve()

    filename = "existencias_productos_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
    return (Path.cwd() / "exports" / filename).resolve()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Exporta un informe .xlsx de productos y existencias desde World Office."
    )
    parser.add_argument("--server", help="Servidor SQL Server. Por defecto WORLD_OFFICE_SERVER.")
    parser.add_argument("--database", help=f"Base de datos. Por defecto {DEFAULT_DATABASE}.")
    parser.add_argument("--username", help="Usuario SQL Server. Por defecto WORLD_OFFICE_USERNAME.")
    parser.add_argument("--password", help="Clave SQL Server. Por defecto WORLD_OFFICE_PASSWORD.")
    parser.add_argument("--driver", help=f"Driver ODBC. Por defecto {DEFAULT_DRIVER}.")
    parser.add_argument("--view", default=DEFAULT_VIEW, help=f"Vista de existencias. Por defecto {DEFAULT_VIEW}.")
    parser.add_argument("--bodega", help="Codigo de bodega para filtrar. Si se omite, exporta todas.")
    parser.add_argument("--output", help="Ruta del archivo .xlsx de salida.")
    parser.add_argument("--sin-detalle-bodega", action="store_true", help="No crea la hoja de detalle por bodega.")
    parser.add_argument("--connection-timeout", type=int, default=int(env("WORLD_OFFICE_CONNECTION_TIMEOUT_SECONDS", "8") or 8))
    parser.add_argument("--query-timeout", type=int, default=int(env("WORLD_OFFICE_QUERY_TIMEOUT_SECONDS", "30") or 30))
    return parser.parse_args()


def main() -> int:
    load_environment()
    args = parse_args()
    output_path = resolve_output_path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    connection = connect(args)
    try:
        cursor = connection.cursor()
        total_rows, metadata = fetch_total_inventory(
            cursor,
            view_name=args.view,
            bodega=args.bodega,
        )
        detail_rows: list[dict[str, object]] = []
        if not args.sin_detalle_bodega:
            detail_rows = fetch_inventory_by_bodega(
                cursor,
                view_name=args.view,
                bodega=args.bodega,
            )
    finally:
        connection.close()

    workbook = create_workbook(
        total_rows=total_rows,
        detail_rows=detail_rows,
        metadata=metadata,
        args=args,
    )
    workbook.save(output_path)

    print(f"Informe generado: {output_path}")
    print(f"Productos exportados: {len(total_rows)}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ReportError as exc:
        print(f"Error: {exc}")
        raise SystemExit(1)
